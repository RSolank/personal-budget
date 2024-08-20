import sys
import os
import warnings
import getopt
import datetime
import openpyxl
from tabulate import tabulate
from uuid import uuid4

EXPECTED_HEADERS = ["Tran Date", "Chq No", "Particulars", "Category", "Sub-Category", "Debit", "Credit", "Balance", "Init. Br"]

def extract_raw_data_from_excel(file_url):
    """
    Extracts transaction data from an excel file and returns a list.
    Handles cases where the file might be in an unexpected format.
    """
    raw_data = []

    try:
        warnings.filterwarnings("ignore", ".*DrawingML*.")
        wb = openpyxl.load_workbook(file_url)
        ws = wb.active

        # Check for expected headers in the first row
        headers = [cell.value for cell in ws[1]]
        if not all(expected_header in headers for expected_header in EXPECTED_HEADERS):
            print(f"Skipping file {file_url}: unexpected headers")
            return None

        # Find the index of the relevant columns based on headers
        date_col_idx = headers.index("Tran Date")
        debit_col_idx = headers.index("Debit")
        credit_col_idx = headers.index("Credit")
        details_col_idx = headers.index("Particulars")

        
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                txn_key = str(uuid4())[:6]  # Unique key to identify individual transaction
                txn_date = row[date_col_idx]
                txn_detail = row[details_col_idx] if len(row) > details_col_idx else None
                debit = row[debit_col_idx] if len(row) > debit_col_idx and row[debit_col_idx] is not None else 0
                credit = row[credit_col_idx] if len(row) > credit_col_idx and row[credit_col_idx] is not None else 0
                if isinstance(txn_date, (datetime.datetime, datetime.date)) and \
                   (isinstance(debit, (int, float)) or isinstance(credit, (int, float))):
                    
                    raw_data.append([txn_key, txn_date, txn_detail, debit, credit])
            except (IndexError, TypeError) as e:
                print(f"Skipping row due to error: {e}")
        
    except Exception as e:
        print(f"Error loading file {file_url}: {e}")

    return raw_data


def process_transactions(txn_data):
    """
    Creates a transactions dictionary from given transaction data.
    """
    transactions = {}

    try:
        for txn in txn_data:
            transactions[txn[0]] = {
                "txn_date": txn[1].strftime("%d-%b-%Y"),
                "txn_detail": txn[2],
                "debit": txn[3],
                "credit": txn[4]
            }

    except Exception as e:
        print(f"Error: {e}")
        return None

    return transactions


def consolidate_transactions_month_wise(transactions):
    """
    Creates a monthly summary from the provided transactions data.
    """
    monthly_summary = {}

    try:
        for txn_key, txn in transactions.items():
            txn_date = datetime.datetime.strptime(txn["txn_date"], "%d-%b-%Y")
            month_key = txn_date.strftime("%m-%Y")
            if month_key not in monthly_summary:
                monthly_summary[month_key] = {
                    "month_name": txn_date.strftime("%B, %Y"),
                    "txn_keys": [txn_key],
                    "debit": txn["debit"],
                    "credit": txn["credit"],
                    "surplus_deficit": txn["credit"] - txn["debit"]
                }
            else:
                monthly_summary[month_key]["txn_keys"].append(txn_key)
                monthly_summary[month_key]["debit"] += txn["debit"]
                monthly_summary[month_key]["credit"] += txn["credit"]
                monthly_summary[month_key]["surplus_deficit"] += (txn["credit"] - txn["debit"])

    except Exception as e:
        print(f"Error: {e}")
        return None

    return monthly_summary


def extract_data_from_source(dir_path):
    """
    Extracts transaction data from files in the source directory and returns it as dictionaries.
    Handles random or malformed files by reporting and skipping them.
    """
    transactions, monthly_summary = {}, {}
    txn_data = []
    
    excel_file_exist = False
    try:
        file_list = os.listdir(dir_path)
        for file_name in file_list:
            if file_name.endswith((".xls", ".xlsx")):
                excel_file_exist = True
                file_url = os.path.join(dir_path, file_name)
                raw_data = extract_raw_data_from_excel(file_url)
                if raw_data:
                    txn_data.extend(raw_data)
                        
        if not excel_file_exist:
            print("No Excel file found in source directory")
            return None, None

        if txn_data:
            transactions = process_transactions(txn_data)
            monthly_summary = consolidate_transactions_month_wise(transactions)
    
    except Exception as e:
        print(f"Error: {e}")
    
    return transactions, monthly_summary


def generate_report_meta():
    """
    Generates meta data for the required report based on user input.
    """
    report_meta = {}
    short_options = "ht:s:o:c:"
    long_options = ["help", "report_type=", "sort_by=", "order_by=", "view_count="]

    try:
        arguments, values = getopt.getopt(sys.argv[1:], short_options, long_options)
    
        for current_argument, current_value in arguments:
            if current_argument in ("-h", "--help"):
                print("Displaying Help: \n")
                with open("README.md") as f:
                    print(f.read())
                return None
                
            elif current_argument in ("-t", "--report_type"):
                report_type = current_value.lower().split("_")
                if "month" in report_type[0]:
                    if "top" in report_type:
                        report_meta["type"] = "MT"
                        report_meta["txn_per_month"] = int(report_type[-1]) if report_type[-1].isnumeric() else 3
                        report_meta["sort_req"] = True
                    else:
                        report_meta["type"] = "M"
                elif "transaction" in report_type[0]:
                    report_meta["type"] = "T"
                    report_meta["view_count"] = report_meta.get("view_count", 20)
                else:
                    report_meta["type"] = "M"
                
            elif current_argument in ("-s", "--sort_by"):
                report_meta["sort_req"] = True
                report_meta["sort_by"] = "D" if "debit" in current_value.lower() else "C" if "credit" in current_value.lower() else "S"

            elif current_argument in ("-o", "--order_by"):
                report_meta["sort_order"] = "D" if any(sub in current_value.lower() for sub in ["descend", "decreas"]) else "A"

            elif current_argument in ("-c", "--view_count"):
                report_meta["view_count"] = int(current_value) if current_value.isnumeric() else 10
        
    except getopt.GetoptError as error:
        print(str(error))
        return None
    
    return report_meta


def generate_report(transactions, monthly_summary, report_meta):
    """
    Generates the report based on the provided transactions data and report metadata.
    """
    def sort_table(table, sort_by_column, descend):
        return sorted(table, key=lambda x: x[sort_by_column], reverse=descend)
    
    def filter_zero_entries(table, column):
        return [row for row in table if row[column] != 0]

    final_report = []
    
    report_type = report_meta.get("type", "M")
    sort_req = report_meta.get("sort_req", False)
    sort_by = report_meta.get("sort_by", "D")
    sort_order = report_meta.get("sort_order", "D")
    
    if report_type == "M":
        table = [[month_detail["month_name"], month_detail["debit"],
                  month_detail["credit"], month_detail["surplus_deficit"]]
                 for month_detail in monthly_summary.values()]
        if sort_req:
            sort_by_column = {"D": 1, "C": 2, "S": 3}.get(sort_by, 1)
            descend = sort_order == "D"
            table = filter_zero_entries(table, sort_by_column)
            final_report = sort_table(table, sort_by_column, descend)
        else:
            final_report = table
    
    elif report_type == "T":
        table = [[transaction["txn_date"], transaction["txn_detail"],
                  transaction["debit"], transaction["credit"]]
                 for transaction in transactions.values()]
        if sort_req:
            sort_by_column = {"D": 2, "C": 3, "S": 0}.get(sort_by, 2)
            descend = sort_order == "D"
            table = filter_zero_entries(table, sort_by_column)
            final_report = sort_table(table, sort_by_column, descend)
        else:
            final_report = table

    else:
        for month_key, month_detail in monthly_summary.items():
            table = [[None, transactions[txn_key]["txn_date"], transactions[txn_key]["txn_detail"],
                      transactions[txn_key]["debit"], transactions[txn_key]["credit"]]
                     for txn_key in month_detail["txn_keys"]]
            sort_by_column = {"D": 3, "C": 4, "S": 1}.get(sort_by, 3)
            descend = sort_order == "D"
            table = filter_zero_entries(table, sort_by_column)
            temp_report = sort_table(table, sort_by_column, descend)
            final_report.append([month_detail["month_name"]])
            final_report.extend(temp_report[:report_meta.get("txn_per_month", 3)])
    
    return final_report


def show_output(report, report_meta):
    """
    Displays the summary of transactions in a tabulated format.
    """
    headers = {
        "M": ["Month", "Debit", "Credit", "Surplus/Deficit"],
        "T": ["Transaction_Date", "Transaction_Details", "Debit", "Credit"],
        "MT": ["Month", "Transaction_Date", "Transaction_Details", "Debit", "Credit"]
    }.get(report_meta.get("type", "M"), ["Month", "Transaction_Date", "Transaction_Details", "Debit", "Credit"])

    view_count = report_meta.get("view_count", len(report))

    if report_meta.get("type", "M") == "MT" and view_count < len(report):
        cnt = 0
        for i in range(len(report)):
            if report[i][0] is not None:
                cnt += 1
            if cnt > view_count:
                view_count = i
                break

    print(tabulate(report[:view_count], headers=headers, tablefmt="grid", numalign="right", floatfmt=".2f"))


if __name__ == "__main__":
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "source_data")
    
    report_meta = generate_report_meta()
    transactions, monthly_summary = extract_data_from_source(dir_path)
    report = generate_report(transactions, monthly_summary, report_meta)
    show_output(report, report_meta)
